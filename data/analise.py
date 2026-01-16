import pandas as pd
import numpy as np
import scipy.stats as stats
from scipy.stats import f_oneway, levene
from scipy.stats import f as f_dist
import os
import warnings
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
import datetime
import locale

# Configurar locale para Português Brasil
try:
    locale.setlocale(locale.LC_ALL, 'pt_BR.UTF-8')
except:
    try:
        locale.setlocale(locale.LC_ALL, 'Portuguese_Brazil.1252')
    except:
        pass

warnings.filterwarnings('ignore')

try:
    from statsmodels.stats.multicomp import pairwise_tukeyhsd
    STATSMODELS_AVAILABLE = True
except ImportError:
    STATSMODELS_AVAILABLE = False

# Funções auxiliares para ANOVA de Welch e Games-Howell
def welch_anova(data_groups):
    """
    Realiza ANOVA de Welch (para variâncias heterogêneas)
    Retorna: F, df1, df2, p_value
    """
    k = len(data_groups)
    
    # Calcular médias e variâncias
    means = [np.mean(g) for g in data_groups]
    vars_ = [np.var(g, ddof=1) for g in data_groups]
    ns = [len(g) for g in data_groups]
    
    # Pesos
    w = [n/v for n, v in zip(ns, vars_)]
    
    # Média ponderada
    mean_w = np.sum([w[i]*means[i] for i in range(k)]) / np.sum(w)
    
    # Estatística F de Welch
    A = np.sum([w[i]*(means[i] - mean_w)**2 for i in range(k)])
    B = 2*(k-2) / (k**2 - 1) * np.sum([(1 - w[i]/np.sum(w))**2 / (ns[i]-1) for i in range(k)])
    
    F = A / (k-1) / (1 + B)
    
    # Graus de liberdade
    df1 = k - 1
    df2 = 1 / (3 * np.sum([(1 - w[i]/np.sum(w))**2 / (ns[i]-1) for i in range(k)]))
    
    # Valor p
    p_value = 1 - f_dist.cdf(F, df1, df2)
    
    return F, df1, df2, p_value

def games_howell(data_groups, group_names):
    """
    Realiza teste post-hoc de Games-Howell para variâncias heterogêneas
    """
    results = []
    k = len(data_groups)
    
    for i in range(k):
        for j in range(i+1, k):
            # Dados dos dois grupos
            data_i = data_groups[i]
            data_j = data_groups[j]
            
            # Estatísticas básicas
            n_i = len(data_i)
            n_j = len(data_j)
            mean_i = np.mean(data_i)
            mean_j = np.mean(data_j)
            var_i = np.var(data_i, ddof=1)
            var_j = np.var(data_j, ddof=1)
            
            # Diferença das médias
            diff = mean_i - mean_j
            
            # Erro padrão
            se = np.sqrt(var_i/n_i + var_j/n_j)
            
            # Graus de liberdade (Welch-Satterthwaite)
            df = (var_i/n_i + var_j/n_j)**2 / ((var_i/n_i)**2/(n_i-1) + (var_j/n_j)**2/(n_j-1))
            
            # Estatística t
            t = diff / se
            
            # Valor p (bicaudal)
            p = 2 * (1 - stats.t.cdf(abs(t), df))
            
            # Intervalo de confiança 95%
            t_crit = stats.t.ppf(0.975, df)
            ci_lower = diff - t_crit * se
            ci_upper = diff + t_crit * se
            
            results.append({
                'Grupo 1': group_names[i],
                'Grupo 2': group_names[j],
                'Média 1': mean_i,
                'Média 2': mean_j,
                'Diferença': diff,
                'Erro Padrão': se,
                'IC 95% Inferior': ci_lower,
                'IC 95% Superior': ci_upper,
                't': t,
                'df': df,
                'p-valor': p,
                'Significativo': p < 0.05,
                'Tipo': 'Games-Howell'
            })
    
    return results

def formatar_numero_br(valor, casas_decimais=4):
    """Formata números no formato brasileiro (vírgula como separador decimal)"""
    if pd.isna(valor):
        return "N/A"
    
    try:
        # Formatar com vírgula como separador decimal
        formato = f"{{:,.{casas_decimais}f}}"
        numero_formatado = formato.format(float(valor))
        # Substituir ponto por vírgula
        return numero_formatado.replace('.', ',')
    except:
        return str(valor)

def formatar_pvalor_br(valor):
    """Formata p-valores no formato brasileiro"""
    if pd.isna(valor):
        return "N/A"
    
    try:
        valor_float = float(valor)
        if valor_float < 0.001:
            return "< 0,001"
        elif valor_float < 0.01:
            return f"{valor_float:.3f}".replace('.', ',')
        else:
            return f"{valor_float:.4f}".replace('.', ',')
    except:
        return str(valor)

def analisar_dataframe_completo(df, nome_arquivo):
    """Analisa um DataFrame e retorna todos os resultados detalhados seguindo o fluxo especificado"""
    resultados_detalhados = []
    resultados_posthoc_completo = []
    
    # Limpeza dos dados
    df = df.dropna(axis=1, how='all')
    df.columns = df.columns.str.strip()
    
    # Identificar colunas numéricas (ignorando identificadores)
    padroes_ignorar = ['episódio', 'episodio', 'episode', 'id', 'codigo', 'código', 'code']
    colunas_numericas = []
    
    for coluna in df.columns:
        coluna_lower = str(coluna).strip().lower()
        ignorar = any(padrao in coluna_lower for padrao in padroes_ignorar)
        
        if not ignorar:
            try:
                dados_convertidos = pd.to_numeric(df[coluna].astype(str).str.replace(',', '.'), errors='coerce')
                if dados_convertidos.notna().sum() / len(dados_convertidos) >= 0.5:
                    colunas_numericas.append(coluna)
            except:
                continue
    
    if not colunas_numericas:
        return {
            'estatisticas': [],
            'comparacoes': [],
            'posthoc': [],
            'correlacoes': []
        }
    
    # Converter apenas as colunas numéricas
    for coluna in colunas_numericas:
        df[coluna] = pd.to_numeric(df[coluna].astype(str).str.replace(',', '.'), errors='coerce')
    
    # Identificar colunas válidas (com pelo menos 3 observações)
    colunas_validas = [col for col in colunas_numericas if len(df[col].dropna()) >= 3]
    
    if len(colunas_validas) < 1:
        return {
            'estatisticas': [],
            'comparacoes': [],
            'posthoc': [],
            'correlacoes': []
        }
    
    # --- SEÇÃO 1: ESTATÍSTICAS DESCRITIVAS COMPLETAS ---
    estatisticas_detalhadas = []
    for coluna in colunas_validas:
        dados = df[coluna].dropna()
        
        # Para amostras grandes (>100), usar intervalo de confiança baseado na distribuição normal
        n = len(dados)
        if n > 0:
            mean_val = dados.mean()
            std_val = dados.std()
            se = std_val / np.sqrt(n)
            
            # Para n > 30, usar distribuição normal; para n <= 30, usar t-student
            if n > 30:
                ci_lower = mean_val - 1.96 * se
                ci_upper = mean_val + 1.96 * se
            else:
                t_crit = stats.t.ppf(0.975, n-1)
                ci_lower = mean_val - t_crit * se
                ci_upper = mean_val + t_crit * se
        else:
            mean_val = std_val = se = ci_lower = ci_upper = np.nan
        
        estatisticas = {
            'Arquivo': nome_arquivo,
            'Variável': coluna,
            'Tipo': 'Estatística Descritiva',
            'Média': mean_val,
            'DP': std_val,
            'Mediana': dados.median(),
            'IC 95% Inferior': ci_lower,
            'IC 95% Superior': ci_upper,
            'Mínimo': dados.min(),
            'Máximo': dados.max(),
            'CV (%)': (std_val / mean_val * 100) if mean_val != 0 else np.nan,
            'N': n
        }
        estatisticas_detalhadas.append(estatisticas)
    
    # --- SEÇÃO 2: COMPARAÇÃO ENTRE GRUPOS (FLUXO ESPECIFICADO) ---
    comparacoes_grupos = []
    
    if len(colunas_validas) >= 2:
        grupos_dados = [df[col].dropna().values for col in colunas_validas]
        N_total = sum(len(g) for g in grupos_dados)
        k = len(grupos_dados)
        
        # TESTE DE LEVENE (Homogeneidade de variâncias)
        try:
            stat_levene, p_levene = levene(*grupos_dados)
            homogeneo = p_levene >= 0.05  # p ≥ 0.05 = homogêneo, p < 0.05 = não homogêneo
            
            comparacoes_grupos.append({
                'Arquivo': nome_arquivo,
                'Variável': 'Todas',
                'Tipo': 'Homogeneidade',
                'Teste': 'Levene',
                'Estatística': stat_levene,
                'p-valor': p_levene,
                'Resultado': 'Homogêneas' if homogeneo else 'Não Homogêneas',
                'Decisão': 'ANOVA Clássica + Tukey' if homogeneo else 'ANOVA Welch + Games-Howell'
            })
            
            # FLUXO DECISIONAL BASEADO NO LEVENE
            if homogeneo:  # p ≥ 0,05
                # ANOVA CLÁSSICA
                try:
                    f_stat, p_anova = f_oneway(*grupos_dados)
                    df_between = k - 1
                    df_within = N_total - k
                    
                    if df_within > 0:
                        comparacoes_grupos.append({
                            'Arquivo': nome_arquivo,
                            'Variável': 'ANOVA',
                            'Tipo': 'Comparação',
                            'Teste': 'ANOVA Clássica',
                            'Estatística': f_stat,
                            'gl entre': df_between,
                            'gl dentro': df_within,
                            'p-valor': p_anova,
                            'Resultado': 'Significativa' if p_anova < 0.05 else 'Não Significativa',
                            'Decisão': 'Prosseguir com Tukey HSD' if p_anova < 0.05 else 'Não necessário'
                        })
                        
                        # TUKEY HSD (apenas se ANOVA for significativa)
                        if p_anova < 0.05 and STATSMODELS_AVAILABLE:
                            try:
                                tukey_data = []
                                tukey_groups = []
                                
                                for i, dados in enumerate(grupos_dados):
                                    tukey_data.extend(dados)
                                    tukey_groups.extend([colunas_validas[i]] * len(dados))
                                
                                tukey_result = pairwise_tukeyhsd(tukey_data, tukey_groups, alpha=0.05)
                                
                                if hasattr(tukey_result, 'summary') and tukey_result.summary() is not None:
                                    for i in range(len(tukey_result.summary().data) - 1):
                                        row = tukey_result.summary().data[i + 1]
                                        if len(row) > 6:
                                            comparacao = {
                                                'Arquivo': nome_arquivo,
                                                'Grupo 1': str(row[0]),
                                                'Grupo 2': str(row[1]),
                                                'Média 1': df[colunas_validas[colunas_validas.index(str(row[0]))]].mean() if str(row[0]) in colunas_validas else np.nan,
                                                'Média 2': df[colunas_validas[colunas_validas.index(str(row[1]))]].mean() if str(row[1]) in colunas_validas else np.nan,
                                                'Diferença': float(row[2]),
                                                'Erro Padrão': float(row[3]),
                                                'IC 95% Inferior': float(row[4]),
                                                'IC 95% Superior': float(row[5]),
                                                'p-valor': float(row[6]) if len(row) > 6 else np.nan,
                                                'Significativo': bool(row[7]) if len(row) > 7 else False,
                                                'Tipo': 'Tukey HSD'
                                            }
                                            resultados_posthoc_completo.append(comparacao)
                            except Exception as e:
                                resultados_posthoc_completo.append({
                                    'Arquivo': nome_arquivo,
                                    'Grupo 1': 'Erro',
                                    'Grupo 2': str(e)[:50],
                                    'Tipo': 'Erro Tukey'
                                })
                except Exception as e:
                    comparacoes_grupos.append({
                        'Arquivo': nome_arquivo,
                        'Variável': 'ANOVA',
                        'Tipo': 'Comparação',
                        'Teste': 'ANOVA Clássica (Erro)',
                        'Estatística': np.nan,
                        'gl entre': np.nan,
                        'gl dentro': np.nan,
                        'p-valor': np.nan,
                        'Resultado': f'Erro: {str(e)[:50]}'
                    })
            
            else:  # p < 0,05
                # ANOVA DE WELCH
                try:
                    F_welch, df1_welch, df2_welch, p_welch = welch_anova(grupos_dados)
                    
                    comparacoes_grupos.append({
                        'Arquivo': nome_arquivo,
                        'Variável': 'ANOVA',
                        'Tipo': 'Comparação',
                        'Teste': 'ANOVA de Welch',
                        'Estatística': F_welch,
                        'gl entre': df1_welch,
                        'gl dentro': df2_welch,
                        'p-valor': p_welch,
                        'Resultado': 'Significativa' if p_welch < 0.05 else 'Não Significativa',
                        'Decisão': 'Prosseguir com Games-Howell' if p_welch < 0.05 else 'Não necessário'
                    })
                    
                    # GAMES-HOWELL (apenas se ANOVA de Welch for significativa)
                    if p_welch < 0.05:
                        try:
                            gh_results = games_howell(grupos_dados, colunas_validas)
                            for gh in gh_results:
                                gh['Arquivo'] = nome_arquivo
                                resultados_posthoc_completo.append(gh)
                        except Exception as e:
                            resultados_posthoc_completo.append({
                                'Arquivo': nome_arquivo,
                                'Grupo 1': 'Erro',
                                'Grupo 2': str(e)[:50],
                                'Tipo': 'Erro Games-Howell'
                            })
                except Exception as e:
                    comparacoes_grupos.append({
                        'Arquivo': nome_arquivo,
                        'Variável': 'ANOVA',
                        'Tipo': 'Comparação',
                        'Teste': 'ANOVA de Welch (Erro)',
                        'Estatística': np.nan,
                        'gl entre': np.nan,
                        'gl dentro': np.nan,
                        'p-valor': np.nan,
                        'Resultado': f'Erro: {str(e)[:50]}'
                    })
        
        except Exception as e:
            comparacoes_grupos.append({
                'Arquivo': nome_arquivo,
                'Variável': 'Todas',
                'Tipo': 'Homogeneidade',
                'Teste': 'Levene',
                'Estatística': np.nan,
                'p-valor': np.nan,
                'Resultado': f'Erro: {str(e)[:50]}'
            })
    
    # --- SEÇÃO 3: MATRIZ DE CORRELAÇÃO (Pearson) ---
    matriz_correlacao = []
    if len(colunas_validas) >= 2:
        # Calcular correlação de Pearson
        df_corr = df[colunas_validas].corr()
        
        for i in range(len(colunas_validas)):
            for j in range(i+1, len(colunas_validas)):
                var1 = colunas_validas[i]
                var2 = colunas_validas[j]
                corr = df_corr.loc[var1, var2]
                
                # Teste de significância da correlação
                n = len(df[[var1, var2]].dropna())
                if n > 2:
                    t_stat = corr * np.sqrt((n-2)/(1-corr**2)) if abs(corr) < 1 else np.nan
                    p_corr = 2 * (1 - stats.t.cdf(abs(t_stat), n-2)) if not pd.isna(t_stat) else np.nan
                else:
                    p_corr = np.nan
                
                matriz_correlacao.append({
                    'Arquivo': nome_arquivo,
                    'Variável 1': var1,
                    'Variável 2': var2,
                    'Correlação (r)': corr,
                    'r²': corr**2,
                    'p-valor': p_corr,
                    'n': n,
                    'Interpretação': interpretar_correlacao(corr),
                    'Tipo': 'Correlação'
                })
    
    return {
        'estatisticas': estatisticas_detalhadas,
        'comparacoes': comparacoes_grupos,
        'posthoc': resultados_posthoc_completo,
        'correlacoes': matriz_correlacao
    }

def interpretar_correlacao(r):
    """Interpreta o valor da correlação"""
    if pd.isna(r):
        return "N/A"
    
    r_abs = abs(r)
    if r_abs >= 0.9:
        return "Muito forte"
    elif r_abs >= 0.7:
        return "Forte"
    elif r_abs >= 0.5:
        return "Moderada"
    elif r_abs >= 0.3:
        return "Fraca"
    elif r_abs >= 0.1:
        return "Muito fraca"
    else:
        return "Desprezível"

def criar_excel_por_arquivo(arquivos_resultados):
    """Cria um Excel com uma aba para cada arquivo CSV"""
    
    wb = Workbook()
    
    # Remover aba padrão
    wb.remove(wb.active)
    
    # Criar aba de índice
    ws_indice = wb.create_sheet(title="ÍNDICE")
    ws_indice.append(["RELATÓRIO DE ANÁLISE ESTATÍSTICA"])
    ws_indice.append([f"Gerado em: {datetime.datetime.now().strftime('%d/%m/%Y %H:%M:%S')}"])
    ws_indice.append([])
    ws_indice.append(["Arquivo", "Variáveis", "Observações", "Teste ANOVA", "p-valor ANOVA", "Teste Post-Hoc", "Aba"])
    
    for nome_arquivo, resultados in arquivos_resultados.items():
        # Criar aba para este arquivo
        nome_aba = nome_arquivo[:31]  # Excel limita a 31 caracteres
        ws = wb.create_sheet(title=nome_aba)
        
        # --- CABEÇALHO DA ABA ---
        ws.append([f"ANÁLISE ESTATÍSTICA: {nome_arquivo}"])
        ws.append([f"Data: {datetime.datetime.now().strftime('%d/%m/%Y %H:%M')}"])
        ws.append([])
        
        # --- 1. ESTATÍSTICAS DESCRITIVAS ---
        ws.append(["1. ESTATÍSTICAS DESCRITIVAS"])
        ws.append(["Variável", "Média", "DP", "Mediana", 
                  "IC 95% Inferior", "IC 95% Superior", "Mínimo", "Máximo", 
                  "CV (%)"])
        
        for estat in resultados['estatisticas']:
            ws.append([
                estat['Variável'],
                formatar_numero_br(estat['Média']),
                formatar_numero_br(estat['DP']),
                formatar_numero_br(estat['Mediana']),
                formatar_numero_br(estat['IC 95% Inferior']),
                formatar_numero_br(estat['IC 95% Superior']),
                formatar_numero_br(estat['Mínimo']),
                formatar_numero_br(estat['Máximo']),
                formatar_numero_br(estat['CV (%)'], 1) if not pd.isna(estat['CV (%)']) else "N/A"
            ])
        
        ws.append([])
        ws.append([])
        
        # --- 2. ANÁLISE DE HOMOGENEIDADE - TESTE DE LEVENE ---
        if resultados['comparacoes']:
            # Encontrar resultado do Levene
            levene_results = [c for c in resultados['comparacoes'] if c.get('Teste') == 'Levene']
            if levene_results:
                ws.append(["2. TESTE DE LEVENE"])
                ws.append(["Estatística F", "p-valor", "Resultado"])
                
                for levene in levene_results:
                    ws.append([
                        formatar_numero_br(levene['Estatística']),
                        formatar_pvalor_br(levene['p-valor']),
                        levene['Resultado']
                    ])
                
                ws.append([])
                ws.append([])
                
                # --- 3. ANOVA ---
                anova_results = [c for c in resultados['comparacoes'] if 'ANOVA' in c.get('Teste', '')]
                for anova in anova_results:
                    if 'Clássica' in anova['Teste']:
                        ws.append(["3. ANOVA CLÁSSICA"])
                    elif 'Welch' in anova['Teste']:
                        ws.append(["3. ANOVA DE WELCH"])
                    else:
                        ws.append(["3. ANOVA"])
                    
                    ws.append(["F", "gl entre", "gl dentro", "p-valor", "Resultado"])
                    ws.append([
                        formatar_numero_br(anova['Estatística']),
                        formatar_numero_br(anova.get('gl entre', np.nan), 0),
                        formatar_numero_br(anova.get('gl dentro', np.nan), 0),
                        formatar_pvalor_br(anova['p-valor']),
                        anova['Resultado']
                    ])
                    
                    ws.append([])
                    ws.append([])
                    
                    # --- 4. TESTE POST-HOC ---
                    if resultados['posthoc']:
                        posthoc_tipo = 'Tukey HSD' if 'Tukey' in str(resultados['posthoc'][0].get('Tipo', '')) else 'Games-Howell'
                        ws.append([f"4. TESTE POST-HOC - {posthoc_tipo}"])
                        
                        # Filtrar apenas resultados válidos
                        valid_posthoc = [p for p in resultados['posthoc'] if 'Erro' not in str(p.get('Tipo', ''))]
                        
                        if valid_posthoc:
                            ws.append(["Grupo 1", "Grupo 2", "Diferença", "Erro Padrão", 
                                      "IC 95% Inferior", "IC 95% Superior", "p-valor", "Significativo"])
                            
                            for posthoc in valid_posthoc:
                                ws.append([
                                    posthoc['Grupo 1'],
                                    posthoc['Grupo 2'],
                                    formatar_numero_br(posthoc.get('Diferença', np.nan)),
                                    formatar_numero_br(posthoc.get('Erro Padrão', np.nan)),
                                    formatar_numero_br(posthoc.get('IC 95% Inferior', np.nan)),
                                    formatar_numero_br(posthoc.get('IC 95% Superior', np.nan)),
                                    formatar_pvalor_br(posthoc.get('p-valor', np.nan)),
                                    "SIM" if posthoc.get('Significativo', False) else "NÃO"
                                ])
                        else:
                            ws.append(["Não foram encontradas comparações post-hoc válidas"])
                        
                        ws.append([])
                        ws.append([])

        # Formatar esta aba
        formatar_aba_excel(ws)
        
        # Adicionar ao índice
        var_count = len(resultados['estatisticas'])
        obs_count = sum(estat['N'] for estat in resultados['estatisticas'])
        
        # Determinar teste ANOVA usado
        teste_anova = "N/A"
        p_anova = "N/A"
        teste_posthoc = "Nenhum"
        
        for comp in resultados['comparacoes']:
            if 'ANOVA' in comp.get('Teste', ''):
                teste_anova = comp['Teste']
                p_anova = formatar_pvalor_br(comp['p-valor'])
                break
        
        if resultados['posthoc']:
            if any('Tukey' in str(p.get('Tipo', '')) for p in resultados['posthoc']):
                teste_posthoc = "Tukey HSD"
            elif any('Games' in str(p.get('Tipo', '')) for p in resultados['posthoc']):
                teste_posthoc = "Games-Howell"
        
        ws_indice.append([
            nome_arquivo,
            var_count,
            obs_count,
            teste_anova,
            p_anova,
            teste_posthoc,
            nome_aba
        ])
    
    # Formatar aba de índice
    formatar_aba_excel(ws_indice)
    
    # Salvar arquivo
    nome_excel = 'Analise_Estatistica.xlsx'
    wb.save(nome_excel)
    
    return nome_excel

def formatar_aba_excel(ws):
    """Aplica formatação personalizada a uma aba - sem sombreamento, sem bordas"""
    # Definir estilos
    fonte_normal = Font(name='Arial', size=10, color='000000')
    fonte_negrito = Font(name='Arial', size=10, bold=True, color='000000')
    fonte_titulo = Font(name='Arial', size=12, bold=True, color='000000')
    
    # Aplicar estilos
    for row in ws.iter_rows():
        for cell in row:
            cell.font = fonte_normal
            
            # Célula vazia ou sem valor
            if cell.value is None:
                continue
            
            # Título principal (linha 1)
            if cell.row == 1:
                cell.font = fonte_titulo
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Títulos das seções (linhas que começam com número)
            elif isinstance(cell.value, str) and cell.column == 1:
                if (cell.value.startswith("1.") or 
                    cell.value.startswith("2.") or 
                    cell.value.startswith("3.") or 
                    cell.value.startswith("4.") or 
                    cell.value.startswith("5.")):
                    cell.font = fonte_negrito
            
            # Cabeçalhos de tabela (linha após título da seção)
            elif cell.row > 1:
                # Verificar se é cabeçalho de tabela
                row_vals = [ws.cell(row=cell.row, column=c).value for c in range(1, ws.max_column + 1)]
                non_empty = sum(1 for val in row_vals if val and str(val).strip())
                
                # Se a linha tem muitos valores (provavelmente é cabeçalho)
                if non_empty > 2 and cell.row < ws.max_row:
                    # Verificar se a linha anterior era título de seção
                    prev_row_val = ws.cell(row=cell.row-1, column=1).value
                    if prev_row_val and any(prev_row_val.startswith(f"{i}.") for i in range(1, 6)):
                        cell.font = fonte_negrito
                        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Ajustar largura das colunas
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        
        for cell in column:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        
        adjusted_width = min(max_length + 2, 30)
        ws.column_dimensions[column_letter].width = adjusted_width

def processar_todos_csv():
    """Processa todos os arquivos CSV na pasta"""
    
    # Listar arquivos CSV
    arquivos_csv = [f for f in os.listdir('.') 
                   if f.lower().endswith('.csv') 
                   and not f.startswith('Analise_Estatistica_')]
    
    if not arquivos_csv:
        print("✗ Nenhum arquivo CSV encontrado na pasta.")
        return
    
    print(f"🔍 Encontrados {len(arquivos_csv)} arquivo(s) CSV")
    print("-" * 50)
    
    arquivos_resultados = {}
    
    for arquivo in arquivos_csv:
        try:
            print(f"📊 Processando: {arquivo}")
            
            # Tentar diferentes encodings
            df = None
            encoding_usado = None
            
            for encoding in ['utf-8', 'latin-1', 'cp1252', 'iso-8859-1', 'utf-8-sig']:
                try:
                    df = pd.read_csv(arquivo, sep=',', decimal=',', encoding=encoding)
                    encoding_usado = encoding
                    break
                except UnicodeDecodeError:
                    continue
                except Exception:
                    continue
            
            if df is None:
                # Última tentativa
                try:
                    df = pd.read_csv(arquivo, sep=',', decimal=',', engine='python')
                    encoding_usado = 'python engine'
                except Exception as e:
                    print(f"  ✗ Erro na leitura: {str(e)[:50]}")
                    continue
            
            print(f"  ✓ Encoding: {encoding_usado}")
            print(f"  ✓ Formato: {df.shape[0]} linhas × {df.shape[1]} colunas")
            
            # Analisar o DataFrame
            resultados = analisar_dataframe_completo(df, arquivo)
            arquivos_resultados[arquivo] = resultados
            
            # Mostrar resumo rápido
            var_count = len(resultados['estatisticas'])
            print(f"  ✓ Variáveis analisadas: {var_count}")
            
            # Mostrar decisão do teste
            for comp in resultados['comparacoes']:
                if comp.get('Teste') == 'Levene':
                    print(f"  ✓ Levene: {comp.get('Resultado', 'N/A')}")
                    print(f"  ✓ Decisão: {comp.get('Decisão', 'N/A')}")
                elif 'ANOVA' in comp.get('Teste', ''):
                    print(f"  ✓ {comp['Teste']}: {comp.get('Resultado', 'N/A')}")
            
            print()
            
        except Exception as e:
            print(f"  ✗ Erro no processamento: {str(e)[:50]}")
            print()
            continue
    
    if arquivos_resultados:
        # Criar Excel com todas as abas
        excel_file = criar_excel_por_arquivo(arquivos_resultados)
        print(f"\n✅ Análise concluída!")
        print(f"📁 Arquivo Excel gerado: {excel_file}")
        print(f"📊 Total de arquivos analisados: {len(arquivos_resultados)}")
        return excel_file
    else:
        print("\n❌ Nenhum arquivo pôde ser analisado.")
        return None

if __name__ == "__main__":
    print("=" * 70)
    print("📈 SISTEMA DE ANÁLISE ESTATÍSTICA")
    print("=" * 70)
    print("FLUXO ESTATÍSTICO:")
    print("1. ESTATÍSTICAS DESCRITIVAS")
    print("2. Teste de Levene (Homogeneidade)")
    print("   - Se p ≥ 0,05 → ANOVA Clássica + Tukey")
    print("   - Se p < 0,05 → ANOVA Welch + Games-Howell")
    print("3. Correlação de Pearson")
    print("\nESTATÍSTICAS PARA N > 100:")
    print("• Intervalo de Confiança baseado na distribuição normal")
    print("• Teste de Levene robusto para grandes amostras")
    print("• ANOVA válida devido ao Teorema do Limite Central")
    print("-" * 50)
    
    print("⚙️  Verificando dependências...")
    
    # Verificar dependências
    try:
        import openpyxl
        print("  ✓ Openpyxl: OK")
    except:
        print("  ✗ Openpyxl não instalado. Instale: pip install openpyxl")
        exit()
    
    if STATSMODELS_AVAILABLE:
        print("  ✓ Statsmodels: OK (Tukey disponível)")
    else:
        print("  ⚠️  Statsmodels não instalado. Tukey não disponível.")
        print("      Instale: pip install statsmodels")
    
    print("\n" + "-" * 50)
    print("Iniciando processamento...\n")
    
    processar_todos_csv()