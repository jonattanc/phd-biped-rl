import pandas as pd
import numpy as np
import scipy.stats as stats
from scipy.stats import f_oneway, shapiro, levene
from scipy.stats import f as f_dist
import os
import warnings
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import datetime

warnings.filterwarnings('ignore')

try:
    from statsmodels.stats.multicomp import pairwise_tukeyhsd
    STATSMODELS_AVAILABLE = True
except ImportError:
    STATSMODELS_AVAILABLE = False

# Funções auxiliares para ANOVA de Welch e Games-Howell
def welch_anova(data_groups):
    """
    Realiza ANOVA de Welch (para variâncias heterogêneas)
    Retorna: F, df1, df2, p_value
    """
    k = len(data_groups)
    
    # Calcular médias e variâncias
    means = [np.mean(g) for g in data_groups]
    vars_ = [np.var(g, ddof=1) for g in data_groups]
    ns = [len(g) for g in data_groups]
    
    # Pesos
    w = [n/v for n, v in zip(ns, vars_)]
    
    # Média ponderada
    mean_w = np.sum([w[i]*means[i] for i in range(k)]) / np.sum(w)
    
    # Estatística F de Welch
    A = np.sum([w[i]*(means[i] - mean_w)**2 for i in range(k)])
    B = 2*(k-2) / (k**2 - 1) * np.sum([(1 - w[i]/np.sum(w))**2 / (ns[i]-1) for i in range(k)])
    
    F = A / (k-1) / (1 + B)
    
    # Graus de liberdade
    df1 = k - 1
    df2 = 1 / (3 * np.sum([(1 - w[i]/np.sum(w))**2 / (ns[i]-1) for i in range(k)]))
    
    # Valor p
    p_value = 1 - f_dist.cdf(F, df1, df2)
    
    return F, df1, df2, p_value

def games_howell(data_groups, group_names):
    """
    Realiza teste post-hoc de Games-Howell para variâncias heterogêneas
    """
    results = []
    k = len(data_groups)
    
    for i in range(k):
        for j in range(i+1, k):
            # Dados dos dois grupos
            data_i = data_groups[i]
            data_j = data_groups[j]
            
            # Estatísticas básicas
            n_i = len(data_i)
            n_j = len(data_j)
            mean_i = np.mean(data_i)
            mean_j = np.mean(data_j)
            var_i = np.var(data_i, ddof=1)
            var_j = np.var(data_j, ddof=1)
            
            # Diferença das médias
            diff = mean_i - mean_j
            
            # Erro padrão
            se = np.sqrt(var_i/n_i + var_j/n_j)
            
            # Graus de liberdade (Welch-Satterthwaite)
            df = (var_i/n_i + var_j/n_j)**2 / ((var_i/n_i)**2/(n_i-1) + (var_j/n_j)**2/(n_j-1))
            
            # Estatística t
            t = diff / se
            
            # Valor p (bicaudal)
            p = 2 * (1 - stats.t.cdf(abs(t), df))
            
            # Intervalo de confiança 95%
            t_crit = stats.t.ppf(0.975, df)
            ci_lower = diff - t_crit * se
            ci_upper = diff + t_crit * se
            
            results.append({
                'Grupo 1': group_names[i],
                'Grupo 2': group_names[j],
                'Média 1': mean_i,
                'Média 2': mean_j,
                'Diferença': diff,
                'Erro Padrão': se,
                'IC 95% Inferior': ci_lower,
                'IC 95% Superior': ci_upper,
                't': t,
                'df': df,
                'p-valor': p,
                'Significativo': p < 0.05,
                'Tipo': 'Games-Howell'
            })
    
    return results

def formatar_valor(valor, formato='geral'):
    """Formata valores para exibição"""
    if pd.isna(valor):
        return "N/A"
    
    if formato == 'pvalor':
        if valor < 0.001:
            return "< 0.001"
        elif valor < 0.01:
            return f"{valor:.3f}"
        else:
            return f"{valor:.4f}"
    elif formato == 'numero':
        return f"{valor:.4f}"
    elif formato == 'inteiro':
        return f"{int(valor)}"
    elif formato == 'porcentagem':
        return f"{valor:.1f}%"
    else:
        return str(valor)

def analisar_dataframe_completo(df, nome_arquivo):
    """Analisa um DataFrame e retorna todos os resultados detalhados"""
    resultados_detalhados = []
    resultados_posthoc_completo = []
    
    # Limpeza dos dados
    df = df.dropna(axis=1, how='all')
    df.columns = df.columns.str.strip()
    
    # Converter dados para numérico
    for coluna in df.columns:
        df[coluna] = pd.to_numeric(df[coluna].astype(str).str.replace(',', '.'), errors='coerce')
    
    # Identificar colunas válidas
    colunas_validas = [col for col in df.columns if len(df[col].dropna()) >= 3]
    
    # --- SEÇÃO 1: ESTATÍSTICAS DESCRITIVAS COMPLETAS ---
    estatisticas_detalhadas = []
    for coluna in colunas_validas:
        dados = df[coluna].dropna()
        
        estatisticas = {
            'Arquivo': nome_arquivo,
            'Variável': coluna,
            'Tipo': 'Estatística Descritiva',
            'Média': dados.mean(),
            'DP': dados.std(),
            'Mediana': dados.median(),
            'IC 95% Inferior': dados.mean() - 1.96 * dados.std() / np.sqrt(len(dados)) if len(dados) > 0 else np.nan,
            'IC 95% Superior': dados.mean() + 1.96 * dados.std() / np.sqrt(len(dados)) if len(dados) > 0 else np.nan,
            'Mínimo': dados.min(),
            'Máximo': dados.max(),
            'CV (%)': (dados.std() / dados.mean() * 100) if dados.mean() != 0 else np.nan,
            'N': len(dados)
        }
        estatisticas_detalhadas.append(estatisticas)
    
    # --- SEÇÃO 2: TESTES DE NORMALIDADE ---
    testes_normalidade = []
    for coluna in colunas_validas:
        dados = df[coluna].dropna()
        if 3 <= len(dados) <= 5000:
            try:
                stat, p_valor = shapiro(dados)
                testes_normalidade.append({
                    'Arquivo': nome_arquivo,
                    'Variável': coluna,
                    'Tipo': 'Normalidade',
                    'Teste': 'Shapiro-Wilk',
                    'Estatística': stat,
                    'p-valor': p_valor,
                    'Resultado': 'Normal' if p_valor > 0.05 else 'Não Normal'
                })
            except:
                pass
    
    # --- SEÇÃO 3: COMPARAÇÃO ENTRE GRUPOS ---
    comparacoes_grupos = []
    if len(colunas_validas) >= 2:
        grupos_dados = [df[col].dropna().values for col in colunas_validas]
        N_total = sum(len(g) for g in grupos_dados)
        k = len(grupos_dados)
        homogeneo = True
        p_levene = 1.0

        # TESTE DE LEVENE
        try:
            stat_levene, p_levene = levene(*grupos_dados)
            homogeneo = p_levene >= 0.05  # p ≥ 0.05 = homogêneo, p < 0.05 = não homogêneo

            comparacoes_grupos.append({
                'Arquivo': nome_arquivo,
                'Variável': 'Todas',
                'Tipo': 'Homogeneidade',
                'Teste': 'Levene',
                'Estatística': stat_levene,
                'p-valor': p_levene,
                'Resultado': 'Homogêneas' if homogeneo else 'Não Homogêneas'
            })
        except Exception as e:
            comparacoes_grupos.append({
                'Arquivo': nome_arquivo,
                'Variável': 'Todas',
                'Tipo': 'Homogeneidade',
                'Teste': 'Levene',
                'Estatística': np.nan,
                'p-valor': np.nan,
                'Resultado': f'Erro: {str(e)[:50]}'
            })
            homogeneo = True  # Default para homogêneo em caso de erro

        # DECISÃO BASEADA EXCLUSIVAMENTE NO RESULTADO DO LEVENE
        # SE p_levene < 0.05 (NÃO HOMOGÊNEO) → SEMPRE usar Welch + Games-Howell
        # SE p_levene ≥ 0.05 (HOMOGÊNEO) → usar ANOVA Clássica + Tukey

        if not homogeneo:  # Variâncias NÃO homogêneas (p < 0.05)
            # --- ANOVA DE WELCH (PARA VARIÂNCIAS HETEROGÊNEAS) ---
            try:
                F_welch, df1_welch, df2_welch, p_welch = welch_anova(grupos_dados)

                comparacoes_grupos.append({
                    'Arquivo': nome_arquivo,
                    'Variável': 'ANOVA',
                    'Tipo': 'Comparação',
                    'Teste': 'ANOVA de Welch',
                    'Estatística': F_welch,
                    'gl entre': df1_welch,
                    'gl dentro': df2_welch,
                    'p-valor': p_welch,
                    'η² (eta)': np.nan,
                    'ω² (omega)': np.nan,
                    'Resultado': 'Significativa' if p_welch < 0.05 else 'Não Significativa',
                    'n': N_total
                })

                # --- GAMES-HOWELL SEMPRE quando usar Welch ---
                try:
                    gh_results = games_howell(grupos_dados, colunas_validas)
                    for gh in gh_results:
                        gh['Arquivo'] = nome_arquivo
                        resultados_posthoc_completo.append(gh)
                except Exception as e:
                    resultados_posthoc_completo.append({
                        'Arquivo': nome_arquivo,
                        'Grupo 1': 'Erro',
                        'Grupo 2': str(e)[:50],
                        'Tipo': 'Erro Games-Howell'
                    })
            except Exception as e:
                comparacoes_grupos.append({
                    'Arquivo': nome_arquivo,
                    'Variável': 'ANOVA',
                    'Tipo': 'Comparação',
                    'Teste': 'ANOVA de Welch (Erro)',
                    'Estatística': np.nan,
                    'gl entre': np.nan,
                    'gl dentro': np.nan,
                    'p-valor': np.nan,
                    'Resultado': f'Erro: {str(e)[:50]}',
                    'n': N_total
                })

        else:  # Variâncias homogêneas (p ≥ 0.05 ou erro no Levene)
            # --- ANOVA CLÁSSICA (PARA VARIÂNCIAS HOMOGÊNEAS) ---
            try:
                f_stat, p_anova = f_oneway(*grupos_dados)
                df_between = k - 1
                df_within = N_total - k

                if df_within > 0:
                    eta2 = (f_stat * df_between) / (f_stat * df_between + df_within)
                    omega2 = max(0, (f_stat - 1) * df_between / (f_stat * df_between + df_within + 1))

                    comparacoes_grupos.append({
                        'Arquivo': nome_arquivo,
                        'Variável': 'ANOVA',
                        'Tipo': 'Comparação',
                        'Teste': 'ANOVA Clássica',
                        'Estatística': f_stat,
                        'gl entre': df_between,
                        'gl dentro': df_within,
                        'p-valor': p_anova,
                        'η² (eta)': eta2,
                        'ω² (omega)': omega2,
                        'Resultado': 'Significativa' if p_anova < 0.05 else 'Não Significativa',
                        'n': N_total
                    })

                    # --- TUKEY HSD SOMENTE quando usar ANOVA Clássica ---
                    if p_anova < 0.05 and STATSMODELS_AVAILABLE:
                        try:
                            tukey_data = []
                            tukey_groups = []

                            for i, dados in enumerate(grupos_dados):
                                tukey_data.extend(dados)
                                tukey_groups.extend([colunas_validas[i]] * len(dados))

                            tukey_result = pairwise_tukeyhsd(tukey_data, tukey_groups, alpha=0.05)

                            # Extrair TODAS as comparações
                            if hasattr(tukey_result, 'summary') and tukey_result.summary() is not None:
                                for i in range(len(tukey_result.summary().data) - 1):
                                    row = tukey_result.summary().data[i + 1]
                                    if len(row) > 6:
                                        comparacao = {
                                            'Arquivo': nome_arquivo,
                                            'Grupo 1': str(row[0]),
                                            'Grupo 2': str(row[1]),
                                            'Média 1': df[colunas_validas[colunas_validas.index(str(row[0]))]].mean() if str(row[0]) in colunas_validas else np.nan,
                                            'Média 2': df[colunas_validas[colunas_validas.index(str(row[1]))]].mean() if str(row[1]) in colunas_validas else np.nan,
                                            'Diferença': float(row[2]),
                                            'Erro Padrão': float(row[3]),
                                            'IC 95% Inferior': float(row[4]),
                                            'IC 95% Superior': float(row[5]),
                                            'p-valor': float(row[6]) if len(row) > 6 else np.nan,
                                            'Significativo': bool(row[7]) if len(row) > 7 else False,
                                            'Tipo': 'Tukey HSD'
                                        }
                                        resultados_posthoc_completo.append(comparacao)
                        except Exception as e:
                            resultados_posthoc_completo.append({
                                'Arquivo': nome_arquivo,
                                'Grupo 1': 'Erro',
                                'Grupo 2': str(e)[:50],
                                'Tipo': 'Erro Tukey'
                            })
                else:
                    comparacoes_grupos.append({
                        'Arquivo': nome_arquivo,
                        'Variável': 'ANOVA',
                        'Tipo': 'Comparação',
                        'Teste': 'ANOVA Clássica',
                        'Estatística': np.nan,
                        'gl entre': df_between,
                        'gl dentro': df_within,
                        'p-valor': np.nan,
                        'η² (eta)': np.nan,
                        'ω² (omega)': np.nan,
                        'Resultado': 'Não Aplicável',
                        'n': N_total
                    })
            except Exception as e:
                comparacoes_grupos.append({
                    'Arquivo': nome_arquivo,
                    'Variável': 'ANOVA',
                    'Tipo': 'Comparação',
                    'Teste': 'ANOVA Clássica (Erro)',
                    'Estatística': np.nan,
                    'gl entre': np.nan,
                    'gl dentro': np.nan,
                    'p-valor': np.nan,
                    'Resultado': f'Erro: {str(e)[:50]}',
                    'n': N_total
                })

        # --- KRUSKAL-WALLIS (SEMPRE calculado, independente de homogeneidade) ---
        try:
            h_stat, p_kw = stats.kruskal(*grupos_dados)
            comparacoes_grupos.append({
                'Arquivo': nome_arquivo,
                'Variável': 'Kruskal-Wallis',
                'Tipo': 'Comparação',
                'Teste': 'Kruskal-Wallis',
                'Estatística': h_stat,
                'gl': k - 1,
                'p-valor': p_kw,
                'Resultado': 'Significativa' if p_kw < 0.05 else 'Não Significativa',
                'n': N_total
            })
        except Exception as e:
            comparacoes_grupos.append({
                'Arquivo': nome_arquivo,
                'Variável': 'Kruskal-Wallis',
                'Tipo': 'Comparação',
                'Teste': 'Kruskal-Wallis',
                'Estatística': np.nan,
                'gl': np.nan,
                'p-valor': np.nan,
                'Resultado': f'Erro: {str(e)[:50]}',
                'n': N_total
            })
    
    # --- SEÇÃO 4: MATRIZ DE CORRELAÇÃO ---
    matriz_correlacao = []
    if len(colunas_validas) >= 2:
        # Calcular correlação de Pearson
        df_corr = df[colunas_validas].corr()
        
        for i in range(len(colunas_validas)):
            for j in range(i+1, len(colunas_validas)):
                var1 = colunas_validas[i]
                var2 = colunas_validas[j]
                corr = df_corr.loc[var1, var2]
                
                # Teste de significância da correlação
                n = len(df[[var1, var2]].dropna())
                if n > 2:
                    t_stat = corr * np.sqrt((n-2)/(1-corr**2)) if abs(corr) < 1 else np.nan
                    p_corr = 2 * (1 - stats.t.cdf(abs(t_stat), n-2)) if not pd.isna(t_stat) else np.nan
                else:
                    p_corr = np.nan
                
                matriz_correlacao.append({
                    'Arquivo': nome_arquivo,
                    'Variável 1': var1,
                    'Variável 2': var2,
                    'Correlação (r)': corr,
                    'r²': corr**2,
                    'p-valor': p_corr,
                    'n': n,
                    'Interpretação': interpretar_correlacao(corr),
                    'Tipo': 'Correlação'
                })
    
    # --- SEÇÃO 5: TESTE T PAR A PAR (se apenas 2 grupos) ---
    testes_t = []
    if len(colunas_validas) == 2:
        var1, var2 = colunas_validas[0], colunas_validas[1]
        dados1 = df[var1].dropna()
        dados2 = df[var2].dropna()
        
        # Teste t para amostras independentes
        t_stat, p_t = stats.ttest_ind(dados1, dados2, equal_var=True)
        
        # Tamanho do efeito (Cohen's d)
        n1, n2 = len(dados1), len(dados2)
        pooled_std = np.sqrt(((n1-1)*dados1.std()**2 + (n2-1)*dados2.std()**2) / (n1+n2-2))
        cohens_d = abs((dados1.mean() - dados2.mean()) / pooled_std) if pooled_std != 0 else np.nan
        
        testes_t.append({
            'Arquivo': nome_arquivo,
            'Variável 1': var1,
            'Variável 2': var2,
            'Teste': 'Teste t',
            'Estatística t': t_stat,
            'gl': n1 + n2 - 2,
            'p-valor': p_t,
            "Cohen's d": cohens_d,
            'Interpretação d': interpretar_cohens_d(cohens_d),
            'Tipo': 'Teste t'
        })
    
    return {
        'estatisticas': estatisticas_detalhadas,
        'normalidade': testes_normalidade,
        'comparacoes': comparacoes_grupos,
        'posthoc': resultados_posthoc_completo,
        'correlacoes': matriz_correlacao,
        'testes_t': testes_t
    }

def interpretar_correlacao(r):
    """Interpreta o valor da correlação"""
    r_abs = abs(r)
    if r_abs >= 0.9:
        return "Muito forte"
    elif r_abs >= 0.7:
        return "Forte"
    elif r_abs >= 0.5:
        return "Moderada"
    elif r_abs >= 0.3:
        return "Fraca"
    elif r_abs >= 0.1:
        return "Muito fraca"
    else:
        return "Desprezível"

def interpretar_cohens_d(d):
    """Interpreta o tamanho do efeito de Cohen"""
    if pd.isna(d):
        return "N/A"
    elif d >= 0.8:
        return "Grande"
    elif d >= 0.5:
        return "Médio"
    elif d >= 0.2:
        return "Pequeno"
    else:
        return "Muito pequeno"

def criar_excel_por_arquivo(arquivos_resultados):
    """Cria um Excel com uma aba para cada arquivo CSV"""
    
    wb = Workbook()
    
    # Remover aba padrão
    wb.remove(wb.active)
    
    # Criar aba de índice
    ws_indice = wb.create_sheet(title="ÍNDICE")
    ws_indice.append(["RELATÓRIO DE ANÁLISE ESTATÍSTICA"])
    ws_indice.append([f"Gerado em: {datetime.datetime.now().strftime('%d/%m/%Y %H:%M:%S')}"])
    ws_indice.append([])
    ws_indice.append(["Arquivo", "Variáveis", "Observações", "ANOVA", "Kruskal-Wallis", "Aba"])
    
    # Dicionário para armazenar resumo
    resumo_arquivos = []
    
    for nome_arquivo, resultados in arquivos_resultados.items():
        # Criar aba para este arquivo
        nome_aba = nome_arquivo[:31]  # Excel limita a 31 caracteres
        ws = wb.create_sheet(title=nome_aba)
        
        # --- CABEÇALHO DA ABA ---
        ws.append([f"ANÁLISE ESTATÍSTICA: {nome_arquivo}"])
        ws.append([f"Data: {datetime.datetime.now().strftime('%d/%m/%Y %H:%M')}"])
        ws.append([])
        
        # --- 1. ESTATÍSTICAS DESCRITIVAS ---
        ws.append(["1. ESTATÍSTICAS DESCRITIVAS"])
        ws.append(["Variável", "Média", "DP", "Mediana", 
                  "IC 95% Inferior", "IC 95% Superior", "Mínimo", "Máximo", 
                  "CV (%)"])
        
        for estat in resultados['estatisticas']:
            ws.append([
                estat['Variável'],
                estat['Média'],
                estat['DP'],
                estat['Mediana'],
                estat['IC 95% Inferior'],
                estat['IC 95% Superior'],
                estat['Mínimo'],
                estat['Máximo'],
                estat['CV (%)']
            ])
        
        ws.append([])
        ws.append([])
        
        # --- 2. TESTES DE NORMALIDADE ---
        if resultados['normalidade']:
            ws.append(["2. TESTES DE NORMALIDADE (Shapiro-Wilk)"])
            ws.append(["Variável", "Estatística W", "p-valor", "Resultado"])
            
            for teste in resultados['normalidade']:
                ws.append([
                    teste['Variável'],
                    teste['Estatística'],
                    teste['p-valor'],
                    teste['Resultado']
                ])
            
            ws.append([])
            ws.append([])
        
        # --- 3. COMPARAÇÕES ENTRE GRUPOS ---
        if resultados['comparacoes']:
            ws.append(["3. COMPARAÇÕES ENTRE GRUPOS"])
            
            for comp in resultados['comparacoes']:
                if comp['Teste'] == 'Levene':
                    ws.append(["Teste de Levene"])
                    ws.append(["Estatística F", "p-valor", "Resultado"])
                    ws.append([
                        comp['Estatística'],
                        comp['p-valor'],
                        comp['Resultado']
                    ])
                    ws.append([])
                
                elif comp['Teste'] == 'ANOVA':
                    ws.append(["ANOVA Paramétrica"])
                    ws.append(["F", "gl entre", "gl dentro", "p-valor", "η²", "ω²", "Resultado"])
                    ws.append([
                        comp['Estatística'],
                        comp['gl entre'],
                        comp['gl dentro'],
                        comp['p-valor'],
                        comp['η² (eta)'],
                        comp['ω² (omega)'],
                        comp['Resultado']
                    ])
                    ws.append([])
                
                elif comp['Teste'] == 'Kruskal-Wallis':
                    ws.append(["Kruskal-Wallis (Não Paramétrico)"])
                    ws.append(["H", "gl", "p-valor", "Resultado"])
                    ws.append([
                        comp['Estatística'],
                        comp['gl'],
                        comp['p-valor'],
                        comp['Resultado']
                    ])
                    ws.append([])
            ws.append([])

        # --- 4. TESTE TUKEY HSD (TODAS COMPARAÇÕES) ---
        if resultados['posthoc'] and any(r['Tipo'] in ['Tukey HSD', 'Games-Howell'] for r in resultados['posthoc']):
            ws.append(["4. TESTE POST-HOC"])
            
            # Determinar qual teste foi usado
            teste_usado = 'Tukey HSD' if any(r['Tipo'] == 'Tukey HSD' for r in resultados['posthoc']) else 'Games-Howell'
            ws.append([f"Teste utilizado: {teste_usado} (baseado no resultado do teste de Levene)"])
            
            ws.append(["Grupo 1", "Grupo 2", "Diferença", "Erro Padrão", 
                      "IC 95% Inferior", "IC 95% Superior", "p-valor", "Significativo"])
            
            posthoc_rows = [r for r in resultados['posthoc'] if r['Tipo'] in ['Tukey HSD', 'Games-Howell']]
            for posthoc in posthoc_rows:
                ws.append([
                    posthoc['Grupo 1'],
                    posthoc['Grupo 2'],
                    posthoc.get('Diferença', np.nan),
                    posthoc.get('Erro Padrão', posthoc.get('Erro Padrão', np.nan)),
                    posthoc.get('IC 95% Inferior', np.nan),
                    posthoc.get('IC 95% Superior', np.nan),
                    posthoc.get('p-valor', posthoc.get('p-valor ajustado', np.nan)),
                    "SIM" if posthoc.get('Significativo', False) else "NÃO"
                ])
            
            ws.append([])
            ws.append([])
        
        # --- 5. CORRELAÇÕES ---
        if resultados['correlacoes']:
            ws.append(["5. MATRIZ DE CORRELAÇÃO (Pearson)"])
            ws.append(["Variável 1", "Variável 2", "r", "r²", "p-valor", "Interpretação"])
            
            for corr in resultados['correlacoes']:
                ws.append([
                    corr['Variável 1'],
                    corr['Variável 2'],
                    corr['Correlação (r)'],
                    corr['r²'],
                    corr['p-valor'],
                    corr['Interpretação']
                ])
            
            ws.append([])
            ws.append([])
        
        # --- 6. TESTE T (se aplicável) ---
        if resultados['testes_t']:
            ws.append(["6. TESTE T PARA AMOSTRAS INDEPENDENTES"])
            ws.append(["Variável 1", "Variável 2", "t", "gl", "p-valor", "Cohen's d", "Interpretação"])
            
            for teste_t in resultados['testes_t']:
                ws.append([
                    teste_t['Variável 1'],
                    teste_t['Variável 2'],
                    teste_t['Estatística t'],
                    teste_t['gl'],
                    teste_t['p-valor'],
                    teste_t["Cohen's d"],
                    teste_t['Interpretação d']
                ])
        
        # Formatar esta aba
        formatar_aba_excel(ws)
        
        # Adicionar ao índice
        var_count = len(resultados['estatisticas'])
        obs_count = sum(estat['N'] for estat in resultados['estatisticas'])
        
        # Determinar teste de variâncias usado
        teste_variancias = "N/A"
        for comp in resultados['comparacoes']:
            if comp.get('Teste') == 'Levene':
                if comp.get('Resultado') == 'Homogêneas':
                    teste_variancias = "Homogêneas (p≥0.05)"
                else:
                    teste_variancias = "Heterogêneas (p<0.05)"
                break
        
        # Determinar teste post-hoc usado
        teste_posthoc = "Nenhum"
        if resultados['posthoc']:
            if any(r['Tipo'] == 'Tukey HSD' for r in resultados['posthoc']):
                teste_posthoc = "Tukey"
            elif any(r['Tipo'] == 'Games-Howell' for r in resultados['posthoc']):
                teste_posthoc = "Games-Howell"
        
        ws_indice.append([
            nome_arquivo,
            var_count,
            obs_count,
            teste_variancias,
            "✓" if any(comp.get('Resultado') == 'Significativa' for comp in resultados['comparacoes'] 
                      if comp.get('Teste') in ['ANOVA Clássica', 'ANOVA de Welch']) else "✗",
            teste_posthoc,
            nome_aba
        ])
    
    # Formatar aba de índice
    formatar_aba_excel(ws_indice)
    
    # Criar aba de resumo consolidado
    criar_aba_resumo(wb, resumo_arquivos)
    
    # Criar aba com todos os resultados posthoc consolidados
    criar_aba_posthoc_consolidado(wb, arquivos_resultados)
    
    # Salvar arquivo
    nome_excel = 'Analise_Estatistica_Individual.xlsx'
    wb.save(nome_excel)
    
    # Aplicar formatação avançada
    formatar_excel_completo(nome_excel)
    
    return nome_excel, resumo_arquivos

def criar_aba_resumo(wb, resumo_arquivos):
    """Cria aba de resumo geral"""
    ws_resumo = wb.create_sheet(title="RESUMO GERAL")
    
    ws_resumo.append(["RESUMO GERAL DA ANÁLISE"])
    ws_resumo.append([f"Data: {datetime.datetime.now().strftime('%d/%m/%Y %H:%M:%S')}"])
    ws_resumo.append([])
    
    # Estatísticas gerais
    total_arquivos = len(resumo_arquivos)
    total_variaveis = sum(r['variaveis'] for r in resumo_arquivos)
    total_observacoes = sum(r['observacoes'] for r in resumo_arquivos)
    arquivos_anova_sig = sum(1 for r in resumo_arquivos if r['anova_sig'])
    arquivos_kw_sig = sum(1 for r in resumo_arquivos if r['kw_sig'])
    
    ws_resumo.append(["ESTATÍSTICAS GERAIS"])
    ws_resumo.append(["Total de Arquivos", total_arquivos])
    ws_resumo.append(["Total de Variáveis", total_variaveis])
    ws_resumo.append(["Total de Observações", total_observacoes])
    ws_resumo.append(["Arquivos com ANOVA Significativa", arquivos_anova_sig])
    ws_resumo.append(["Arquivos com Kruskal-Wallis Significativo", arquivos_kw_sig])
    ws_resumo.append([])
    
    # Tabela detalhada
    ws_resumo.append(["DETALHAMENTO POR ARQUIVO"])
    ws_resumo.append(["Arquivo", "Variáveis", "Observações", "ANOVA Sig.", "K-W Sig."])
    
    for resumo in resumo_arquivos:
        ws_resumo.append([
            resumo['arquivo'],
            resumo['variaveis'],
            resumo['observacoes'],
            "✓" if resumo['anova_sig'] else "✗",
            "✓" if resumo['kw_sig'] else "✗"
        ])
    
    ws_resumo.append([])
    ws_resumo.append(["LEGENDA"])
    ws_resumo.append(["✓ = Significativo (p < 0.05)", "✗ = Não Significativo"])
    
    formatar_aba_excel(ws_resumo)

def criar_aba_posthoc_consolidado(wb, arquivos_resultados):
    """Cria aba com todos os resultados posthoc consolidados"""
    ws_posthoc = wb.create_sheet(title="POSTHOC CONSOLIDADO")
    
    ws_posthoc.append(["RESULTADOS POST-HOC CONSOLIDADOS (Tukey HSD e Games-Howell)"])
    ws_posthoc.append([f"Data: {datetime.datetime.now().strftime('%d/%m/%Y %H:%M:%S')}"])
    ws_posthoc.append([])
    ws_posthoc.append(["Arquivo", "Teste", "Grupo 1", "Grupo 2", "Diferença", "Erro Padrão", 
                      "IC 95% Inferior", "IC 95% Superior", "p-valor", "Significativo"])
    
    # Coletar todos os resultados posthoc
    todos_posthoc = []
    for nome_arquivo, resultados in arquivos_resultados.items():
        for posthoc in resultados['posthoc']:
            if posthoc['Tipo'] in ['Tukey HSD', 'Games-Howell']:
                posthoc['Arquivo'] = nome_arquivo
                todos_posthoc.append(posthoc)
    
    # Ordenar por significância
    todos_posthoc.sort(key=lambda x: (not x.get('Significativo', False), 
                                    x.get('p-valor', 1)))
    
    for posthoc in todos_posthoc:
        ws_posthoc.append([
            posthoc['Arquivo'],
            posthoc['Tipo'],
            posthoc['Grupo 1'],
            posthoc['Grupo 2'],
            posthoc.get('Diferença', np.nan),
            posthoc.get('Erro Padrão', posthoc.get('Erro Padrão', np.nan)),
            posthoc.get('IC 95% Inferior', np.nan),
            posthoc.get('IC 95% Superior', np.nan),
            posthoc.get('p-valor', posthoc.get('p-valor ajustado', np.nan)),
            "SIM" if posthoc.get('Significativo', False) else "NÃO"
        ])
    
    formatar_aba_excel(ws_posthoc)

def formatar_aba_excel(ws):
    """Aplica formatação básica a uma aba"""
    # Definir estilos
    estilo_titulo = Font(name='Calibri', size=14, bold=True, color='FFFFFF')
    fill_titulo = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    
    estilo_subtitulo = Font(name='Calibri', size=12, bold=True, color='FFFFFF')
    fill_subtitulo = PatternFill(start_color='4F81BD', end_color='4F81BD', fill_type='solid')
    
    estilo_cabecalho = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    fill_cabecalho = PatternFill(start_color='95B3D7', end_color='95B3D7', fill_type='solid')
    
    estilo_borda = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )
    
    # Aplicar estilos
    for row in ws.iter_rows():
        for cell in row:
            cell.border = estilo_borda
            
            # Verificar se é título (linha 1)
            if cell.row == 1 and cell.value:
                cell.font = estilo_titulo
                cell.fill = fill_titulo
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Verificar se é subtítulo (linhas com ESTATÍSTICAS, TESTES, etc.)
            elif cell.value and any(keyword in str(cell.value) for keyword in 
                                   ['ESTATÍSTICAS', 'TESTES', 'COMPARAÇÕES', 'TUKEY', 'CORRELAÇÃO', 'RESUMO']):
                if cell.column == 1:  # Apenas primeira coluna
                    cell.font = estilo_subtitulo
                    cell.fill = fill_subtitulo
            
            # Verificar se é cabeçalho de tabela (primeira linha após título)
            elif cell.value and cell.row > 1:
                # Verificar se esta linha tem muitos valores não vazios (provavelmente cabeçalho)
                row_vals = [ws.cell(row=cell.row, column=c).value for c in range(1, ws.max_column + 1)]
                if sum(1 for val in row_vals if val and str(val).strip()) > 3:
                    cell.font = estilo_cabecalho
                    cell.fill = fill_cabecalho
    
    # Ajustar largura das colunas
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        
        for cell in column:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width

def formatar_excel_completo(nome_arquivo):
    """Aplica formatação avançada ao Excel"""
    wb = load_workbook(nome_arquivo)
    
    for ws in wb.worksheets:
        # Adicionar formatação condicional para valores significativos
        if ws.max_row > 10:  # Apenas em abas com dados
            # Formatar células com ✓ em verde
            green_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
            green_font = Font(color='006100')
            
            # Formatar células com ✗ em vermelho claro
            red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
            red_font = Font(color='9C0006')
            
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
                for cell in row:
                    if cell.value and isinstance(cell.value, str):
                        if '✓' in cell.value or 'SIM' in cell.value:
                            cell.fill = green_fill
                            cell.font = green_font
                        elif '✗' in cell.value or 'NÃO' in cell.value or 'Erro' in cell.value:
                            cell.fill = red_fill
                            cell.font = red_font
                        elif 'Significativa' in cell.value:
                            cell.fill = green_fill
                            cell.font = green_font
                        elif 'Não Significativa' in cell.value or 'Não Normal' in cell.value or 'Não Homogêneas' in cell.value:
                            cell.fill = red_fill
                            cell.font = red_font
    
    wb.save(nome_arquivo)

def processar_todos_csv():
    """Processa todos os arquivos CSV na pasta"""
    
    # Listar arquivos CSV
    arquivos_csv = [f for f in os.listdir('.') 
                   if f.lower().endswith('.csv') 
                   and not f.startswith('Analise_Estatistica_')]
    
    if not arquivos_csv:
        print("✗ Nenhum arquivo CSV encontrado na pasta.")
        return
    
    print(f"🔍 Encontrados {len(arquivos_csv)} arquivo(s) CSV")
    print("-" * 50)
    
    arquivos_resultados = {}
    
    for arquivo in arquivos_csv:
        try:
            print(f"📊 Processando: {arquivo}")
            
            # Tentar diferentes encodings
            df = None
            encoding_usado = None
            
            for encoding in ['utf-8', 'latin-1', 'cp1252', 'iso-8859-1', 'utf-8-sig']:
                try:
                    df = pd.read_csv(arquivo, sep=',', decimal=',', encoding=encoding)
                    encoding_usado = encoding
                    break
                except UnicodeDecodeError:
                    continue
                except Exception:
                    continue
            
            if df is None:
                # Última tentativa
                try:
                    df = pd.read_csv(arquivo, sep=',', decimal=',', engine='python')
                    encoding_usado = 'python engine'
                except Exception as e:
                    print(f"  ✗ Erro na leitura: {str(e)[:50]}")
                    continue
            
            print(f"  ✓ Encoding: {encoding_usado}")
            print(f"  ✓ Formato: {df.shape[0]} linhas × {df.shape[1]} colunas")
            
            # Analisar o DataFrame
            resultados = analisar_dataframe_completo(df, arquivo)
            arquivos_resultados[arquivo] = resultados
            
            # Mostrar resumo rápido
            var_count = len(resultados['estatisticas'])
            tukey_count = len([r for r in resultados['posthoc'] if r['Tipo'] == 'Tukey HSD'])
            
            print(f"  ✓ Variáveis analisadas: {var_count}")
            if tukey_count > 0:
                print(f"  ✓ Comparações posthoc: {tukey_count}")
            print()
            
        except Exception as e:
            print(f"  ✗ Erro no processamento: {str(e)[:50]}")
            import traceback
            traceback.print_exc()
            print()
            continue
    
    if arquivos_resultados:
        # Criar Excel com todas as abas
        excel_file, resumo_arquivos = criar_excel_por_arquivo(arquivos_resultados)

if __name__ == "__main__":
    print("=" * 70)
    print("📈 SISTEMA DE ANÁLISE ESTATÍSTICA AVANÇADA")
    print("=" * 70)
    print("Versão: 1.0 | Uma aba por arquivo CSV | Tukey completo")
    print("\n⚙️  Verificando dependências...")
    
    # Verificar dependências
    dependencias = {
        'pandas': '✓',
        'numpy': '✓',
        'scipy': '✓',
        'openpyxl': '✗',
        'statsmodels': '✗'
    }
    
    try:
        import openpyxl
        dependencias['openpyxl'] = '✓'
    except:
        print("  ⚠️  Openpyxl não instalado. Instale: pip install openpyxl")
    
    if STATSMODELS_AVAILABLE:
        dependencias['statsmodels'] = '✓'
    else:
        print("  ⚠️  Statsmodels não instalado. Tukey não disponível.")
        print("      Instale: pip install statsmodels")
    
    print("\n" + "-" * 50)
    print("Iniciando processamento...\n")
    
    processar_todos_csv()